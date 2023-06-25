
import requests
import json
import jsonpath
from configparser import ConfigParser
import openpyxl



config = ConfigParser()
config.read("/home/administrateur/PycharmProjects/TestAutomation/InputFiles/Config.cfg")
base_url = config.get('APIAutomation', 'base_url')


def test_add_single_students():
    api_url = f'{base_url}/api/studentsDetails'
    add_student_file = "/home/administrateur/Documents/API_Automation/add_student.json"
    with open(add_student_file, 'r', encoding='utf-8', newline='') as json_file:
        add_student_json = json.loads(json_file.read())
        post_response = requests.post(api_url, add_student_json)
    assert post_response.status_code == 201


def test_add_multiple_student():
    api_url = f'{base_url}/api/studentsDetails'
    wb = openpyxl.load_workbook('/home/administrateur/Documents/API_Automation/students_excel.xlsx')
    working_sheet = wb['Sheet1']
    rows = working_sheet.max_row  # Calculate the number of rows
    add_student_file = "/home/administrateur/Documents/API_Automation/add_student.json"
    with open(add_student_file, 'r', encoding='utf-8', newline='') as json_file:
        add_student_json_request = json.loads(json_file.read())
    for i in range(2, rows + 1):
        cell_first_name = working_sheet.cell(row=i, column=1)
        cell_middle_name = working_sheet.cell(row=i, column=2)
        cell_last_name = working_sheet.cell(row=i, column=3)
        cell_date_of_birth = working_sheet.cell(row=i, column=4)
        add_student_json_request['first_name'] = cell_first_name.value  # Adding cell_first_name value to json
        add_student_json_request['middle_name'] = cell_middle_name.value  # Adding cell_middle_name value to json
        add_student_json_request['last_name'] = cell_last_name.value  # Adding cell_last_name value to json
        add_student_json_request['date_of_birth'] = cell_date_of_birth.value  # Adding cell_date_of_birth value to json
        response = requests.post(api_url, add_student_json_request)
        print(response.status_code)
        assert response.status_code == 201